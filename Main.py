from StockClass import Stock
from sedScraper import Scraper
from openpyxl import *
from ErrorChecker import ErrorChecker
from Organizer import Organizer


def takeRandomSample(): # used for testing the stock class
    stockList = []
    row = 3
    while(sheet["A"+str(row)].value!=0):
        if(sheet["L"+str(row)].value == 1):
            newStock = Stock(sheet["A"+str(row)].value, sheet["B"+str(row)].value, sheet["C"+str(row)].value,
            sheet["D"+str(row)].value, sheet["E"+str(row)].value, sheet["F"+str(row)].value,
            sheet["G"+str(row)].value, sheet["H"+str(row)].value, sheet["I"+str(row)].value,
            sheet["L"+str(row)].value, sheet["M"+str(row)].value)
            stockList.append(newStock)
        row += 1
    return stockList

def takeAll():  # returns a list of all the stocks from the excel file
    stockList = []
    row = 2
    while(sheet.cell(row, 7).value!=0):
        print(row)
        newStock = Stock(
            sheet.cell(row, 1).value, 
        sheet.cell(row, 2).value, 
        sheet.cell(row, 3).value,
        sheet.cell(row, 4).value, 
        sheet.cell(row, 5).value, 
        sheet.cell(row, 6).value,
        sheet.cell(row, 7).value, 
        sheet.cell(row, 8).value, 
        sheet.cell(row, 9).value,
        )
        stockList.append(newStock)
        row += 1
    return stockList

def printStockList(StockList): # prints the names of the stocks taken in
    for i in range(len(stockList)):
        print("$"+ stockList[i].return_ipo_ticker() + " " + stockList[i].return_ipo_name()
        + " " + str(stockList[i].return_ipo_year()) + " " + str(stockList[i].return_f10k5())
        + " " + str(stockList[i].return_f10k10()))

def get10ksIPONAME(StockList): # downloads the 10ks for the stocks using IPO name
    scraper = Scraper()
    print("DOWNLOADING STONKS")
    for i in range(len(stockList)):
        # print("$"+ stockList[i].return_ipo_ticker() + " " + stockList[i].return_ipo_name()
        # + " "+str(stockList[i].return_ipo_year())+ " "+str(stockList[i].return_f10k5())
        # + " "+str(stockList[i].return_f10k10()))
        scraper.get10kyear(stockList[i].return_ipo_name(), stockList[i].return_ipo_year())
        scraper.get10kyear(stockList[i].return_ipo_name(), stockList[i].return_f10k5())
        scraper.get10kyear(stockList[i].return_ipo_name(), stockList[i].return_f10k10())
        if (i%20 == 0):
            print(i)
    print("DONE DOWNLOADING STONKS")


workbook = load_workbook(filename="StocksInfo.xlsx", read_only= True) # setting up excel file
print(workbook.sheetnames)
sheet = workbook.active
print(sheet)
stockList = []
stockList = takeAll()
printStockList(stockList)
get10ksIPONAME(stockList)
missedStocks = ErrorChecker.checkMissing(stockList, './sec-edgar-filings')
for i in missedStocks:
    print(i.return_ipo_name())
print(len(missedStocks))
Organizer.rename('./TestFiles/')  # renames downloaded files to include company name
Organizer.moveFiles('./testFiles/', './TestMove/') # moves downloaded files to one director
